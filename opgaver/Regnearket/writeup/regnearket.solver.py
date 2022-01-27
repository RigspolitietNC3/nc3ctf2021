# Bruger: https://openpyxl.readthedocs.io/en/stable/
from openpyxl import load_workbook
from PIL import Image


wb = load_workbook(filename = 'nc3ctf2021_regnearket.xlsx', read_only=True)
ws = wb['Sheet2']

img = Image.new( 'RGB', (ws.max_row, ws.max_column), "black") # Create a new black image
pixels = img.load()

x = 0
y = 0
for row in ws.iter_rows() :
    for cell in row:
        cellValue = cell.value
        # =xor(1848711178, 1851994913)
        cellValue = cellValue[5:]
        i1 = cellValue.find(',')
        i2 = cellValue.find(')')
        pixelAsInt = int(cellValue[:i1])
        xorKey = int(cellValue[i1 + 1:i2])
        pixelAsInt = pixelAsInt ^ xorKey
        r = (pixelAsInt & 0xFF0000) >> 16
        g = (pixelAsInt & 0x00FF00) >> 8
        b = (pixelAsInt & 0x0000FF)
        pixels[y, x] = (r, g ,b)
        x += 1

    y += 1
    x = 0

img.save('regnearket.extract.png')
